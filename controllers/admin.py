from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_file
from flask_login import login_required, current_user
from models import User, Course, Group, Category, Certificate, CertificateType, Announcement, PasswordReset, Progress, Video, db
from services.course_service import CourseService
from utils.decorators import admin_required
import uuid
from datetime import datetime, timedelta
import os
import shutil
import sqlite3
import json
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import secrets

admin_bp = Blueprint('admin', __name__)

@admin_bp.route('/courses')
@login_required
@admin_required
def courses():
    """List all courses"""
    courses = Course.query.all()
    return render_template('admin_courses.html', courses=courses)

@admin_bp.route('/users')
@login_required
@admin_required
def users():
    """List all users"""
    users = User.query.filter_by(is_admin=False).all()
    return render_template('admin_users.html', users=users)

@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    """Edit a user"""
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        user.first_name = request.form.get('first_name')
        user.last_name = request.form.get('last_name')
        user.email = request.form.get('email')
        user.is_admin = True if request.form.get('is_admin') == 'on' else False
        db.session.commit()
        flash('Kullanıcı bilgileri güncellendi.', 'success')
        return redirect(url_for('admin_users'))
    return render_template('edit_user.html', user=user)

@admin_bp.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    """Delete a user"""
    user = User.query.get_or_404(user_id)
    if user.is_admin:
        flash('Admin kullanıcılar silinemez.', 'danger')
        return redirect(url_for('admin_users'))
    # Kullanıcıya ait tüm ilerlemeleri sil
    Progress.query.filter_by(user_id=user.id).delete()
    db.session.delete(user)
    db.session.commit()
    flash('Kullanıcı başarıyla silindi.', 'success')
    return redirect(url_for('admin_users'))

@admin_bp.route('/reports')
@login_required
@admin_required
def reports():
    """Admin reports page"""
    courses = Course.query.all()
    return render_template('admin_reports.html', courses=courses, now=datetime.now)

@admin_bp.route('/reports/download', methods=['POST'])
@login_required
@admin_required
def download_multi_report():
    """
    Çoklu kurs raporunu indir (Excel)
    ---
    tags:
      - Admin
    consumes:
      - application/x-www-form-urlencoded
    parameters:
      - name: course_ids
        in: formData
        type: array
        items:
          type: integer
        required: true
        description: Kurs ID listesi
    produces:
      - application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    responses:
      200:
        description: Excel dosyası
        schema:
          type: file
      302:
        description: Yönlendirme (HTML response)
    """
    course_ids = request.form.getlist('course_ids')
    if not course_ids:
        flash('En az bir kurs seçmelisiniz.')
        return redirect(url_for('admin_reports'))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Raporlar'
    row = 1
    
    for course_id in course_ids:
        course = Course.query.get(int(course_id))
        # Kurs başlığı
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value=f"{course.title} - Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=13, color='4f8cff')
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        
        # Sütun başlıkları
        headers = ['Kullanıcı Adı', 'E-posta', 'T. İçerik', 'Tamamlanan', 'İlerleme (%)', 'Test Sonucu', 'Durum', 'Son Tamamlanma']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='6A82FB', end_color='6A82FB', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin', color='B4B4B4'), right=Side(style='thin', color='B4B4B4'), top=Side(style='thin', color='B4B4B4'), bottom=Side(style='thin', color='B4B4B4'))
        row += 1
        
        # Kullanıcı satırları
        for user in course.assigned_users:
            # Yeni progress sistemi ile hesapla
            progress_details = course.get_user_progress(user)
            
            # Test sonucu ve durumu
            test_score = '-'
            test_status = 'Devam Ediyor'
            last_completed_at = ''
            
            if progress_details.passed_test:
                test_score = progress_details.test_score if progress_details.test_score is not None else '-'
                test_status = 'Tamamlandı'
            elif progress_details.is_completed:
                test_status = 'Tamamlandı'
            
            # Son tamamlanma tarihini bul
            latest_date = None
            for video in course.videos:
                progress = Progress.query.filter_by(user_id=user.id, video_id=video.id).first()
                if progress and progress.completed and progress.completed_at:
                    if latest_date is None or progress.completed_at > latest_date:
                        latest_date = progress.completed_at
            
            if latest_date:
                last_completed_at = latest_date.strftime('%d.%m.%Y %H:%M')
            
            ws.cell(row=row, column=1, value=f"{user.first_name} {user.last_name}")
            ws.cell(row=row, column=2, value=user.email)
            ws.cell(row=row, column=3, value=progress_details.total_steps)
            ws.cell(row=row, column=4, value=progress_details.completed_steps)
            ws.cell(row=row, column=5, value=f"{progress_details.progress_percent}%")
            ws.cell(row=row, column=6, value=test_score)
            ws.cell(row=row, column=7, value=test_status)
            ws.cell(row=row, column=8, value=last_completed_at)
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = Border(left=Side(style='thin', color='B4B4B4'), right=Side(style='thin', color='B4B4B4'), top=Side(style='thin', color='B4B4B4'), bottom=Side(style='thin', color='B4B4B4'))
            row += 1
        # Her kursun tablosundan sonra bir satır boşluk bırak
        row += 1
    
    # Sütun genişlikleri
    for col in ws.columns:
        first_cell = next((cell for cell in col if hasattr(cell, 'column_letter')), None)
        if not first_cell:
            continue
        column = first_cell.column_letter
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 4
    
    report_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'multi_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    wb.save(report_path)
    return send_file(report_path, as_attachment=True)

@admin_bp.route('/new-course', methods=['GET', 'POST'])
@login_required
@admin_required
def new_course():
    """Create a new course"""
    if request.method == 'POST':
        # Eski CourseService ile kurs ve içerik kaydı
        new_course, error = CourseService.create_course(request.form, request.files)
        if error:
            flash(f'Kurs oluşturulamadı: {error}', 'danger')
            users = User.query.filter_by(is_admin=False).all()
            groups = Group.query.all()
            categories = Category.query.all()
            return render_template('new_course.html', users=users, groups=groups, categories=categories)
        flash('Kurs ve içerikler başarıyla oluşturuldu.', 'success')
        return redirect(url_for('admin.courses'))
    users = User.query.filter_by(is_admin=False).all()
    groups = Group.query.all()
    categories = Category.query.all()
    return render_template('new_course.html', users=users, groups=groups, categories=categories)

@admin_bp.route('/groups', methods=['GET', 'POST'])
@login_required
@admin_required
def groups():
    """Manage groups"""
    if request.method == 'POST':
        group_name = request.form.get('group_name')
        if group_name:
            new_group = Group(name=group_name)
            db.session.add(new_group)
            db.session.commit()
            flash('Grup başarıyla eklendi!', 'success')
            return redirect(url_for('admin_groups'))
    
    groups = Group.query.all()
    return render_template('admin_groups.html', groups=groups)

@admin_bp.route('/groups/<int:group_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_group(group_id):
    """Delete a group"""
    group = Group.query.get_or_404(group_id)
    db.session.delete(group)
    db.session.commit()
    flash('Grup silindi.', 'success')
    return redirect(url_for('admin_groups'))

@admin_bp.route('/groups/<int:group_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_group(group_id):
    """Edit a group"""
    group = Group.query.get_or_404(group_id)
    users = User.query.filter_by(is_admin=False).all()
    if request.method == 'POST':
        group.name = request.form.get('group_name')
        user_ids = request.form.getlist('user_ids')
        group.users = User.query.filter(User.id.in_(user_ids)).all()
        db.session.commit()
        flash('Grup güncellendi.', 'success')
        return redirect(url_for('admin_groups'))
    return render_template('edit_group.html', group=group, users=users)

@admin_bp.route('/announcements', methods=['GET', 'POST'])
@login_required
@admin_required
def announcements():
    """Manage announcements"""
    if request.method == 'POST':
        title = request.form.get('title')
        content = request.form.get('content')
        if title and content:
            ann = Announcement(title=title, content=content, author_id=current_user.id)
            db.session.add(ann)
            db.session.commit()
            flash('Duyuru eklendi.', 'success')
        else:
            flash('Başlık ve içerik zorunlu.', 'danger')
    announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
    return render_template('admin_announcements.html', announcements=announcements)

@admin_bp.route('/announcements/delete/<int:ann_id>', methods=['POST'])
@login_required
@admin_required
def delete_announcement(ann_id):
    ann = Announcement.query.get_or_404(ann_id)
    db.session.delete(ann)
    db.session.commit()
    flash('Duyuru silindi.', 'success')
    return redirect(url_for('admin.announcements'))

@admin_bp.route('/certificates')
@login_required
@admin_required
def certificates():
    """List all certificates"""
    certificates = Certificate.query.order_by(Certificate.issue_date.desc()).all()
    return render_template('admin_certificates.html', certificates=certificates)

@admin_bp.route('/certificate-operations', methods=['GET', 'POST'])
@login_required
@admin_required
def certificate_operations():
    """Certificate operations"""
    courses = Course.query.all()
    certificate_types = CertificateType.query.all()
    course_id = request.args.get('course_id', type=int)
    selected_course = Course.query.get(course_id) if course_id else None
    
    eligible_users = []
    if selected_course:
        for user in selected_course.assigned_users:
            # Kullanıcının bu kurs için zaten bir sertifikası olup olmadığını kontrol et
            existing_cert = Certificate.query.filter(
                Certificate.user_id == user.id,
                Certificate.course_id == selected_course.id
            ).first()

            if not existing_cert:
                user_progress = selected_course.get_user_progress(user)
                # Sadece kursu ve varsa testi geçen kullanıcıları ekle
                if (selected_course.test_required and user_progress.passed_test) or (not selected_course.test_required and user_progress.all_content_completed):
                    eligible_users.append({'user': user, 'progress': user_progress})

    if request.method == 'POST':
        user_ids = request.form.getlist('user_ids')
        certificate_file = request.files.get('certificate_file')

        # Belge türü zorunlu kontrolü kaldırıldı
        # Sertifika dosyası yoksa otomatik PDF oluştur
        filename = None
        if certificate_file and certificate_file.filename:
            filename = secure_filename(f"cert_{datetime.utcnow().timestamp()}_{certificate_file.filename}")
            upload_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            certificate_file.save(upload_path)
        # Otomatik PDF oluşturma
        elif selected_course and user_ids:
            from utils.helpers import generate_certificate_pdf
            for user_id in user_ids:
                user = User.query.get(user_id)
                if user:
                    certificate_number = f"CERT-{uuid.uuid4().hex[:8].upper()}"
                    issued_at = datetime.utcnow()
                    filename = f"auto_cert_{user.id}_{selected_course.id}_{issued_at.strftime('%Y%m%d%H%M%S')}.pdf"
                    upload_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
                    user_fullname = f"{user.first_name} {user.last_name}"
                    course_title = selected_course.title
                    progress = selected_course.get_user_progress(user)
                    score = progress.test_score if hasattr(progress, 'test_score') else None
                    date_str = issued_at.strftime('%d.%m.%Y')
                    generate_certificate_pdf(user_fullname, course_title, score, date_str, upload_path, certificate_number=certificate_number)
                    certificate = Certificate(
                        user_id=user.id,
                        course_id=selected_course.id,
                        certificate_path=filename,
                        certificate_number=certificate_number,
                        issued_at=issued_at
                    )
                    db.session.add(certificate)
            db.session.commit()
            flash(f'{len(user_ids)} kullanıcıya otomatik sertifika atandı.', 'success')
            return redirect(url_for('admin_certificate_operations', course_id=course_id))
        # Normal dosya yükleme ile devam
        else:
            for user_id in user_ids:
                user = User.query.get(user_id)
                if user:
                    certificate = Certificate(
                        user_id=user.id,
                        course_id=selected_course.id,
                        certificate_path=filename
                    )
                    db.session.add(certificate)
            db.session.commit()
            flash(f'{len(user_ids)} kullanıcıya başarıyla sertifika atandı.', 'success')
            return redirect(url_for('admin_certificate_operations', course_id=course_id))
    
    return render_template('admin_certificate_operations.html', 
                           courses=courses, 
                           selected_course=selected_course,
                           eligible_users=eligible_users,
                           certificate_types=certificate_types,
                           course_id=course_id)

@admin_bp.route('/user-certificates', methods=['GET', 'POST'])
@login_required
@admin_required
def user_certificates():
    """Manage user certificates"""
    users = User.query.filter_by(is_admin=False).all()
    selected_user_id = request.args.get('user_id', type=int)
    selected_user = User.query.get(selected_user_id) if selected_user_id else None
    user_certificates = []
    if selected_user:
        user_certificates = Certificate.query.filter_by(user_id=selected_user.id).all()
    return render_template('admin_user_certificates.html', users=users, selected_user=selected_user, user_certificates=user_certificates)

@admin_bp.route('/dashboard', methods=['GET'])
@login_required
@admin_required
def admin_dashboard():
    courses = Course.query.all()
    course_id = request.args.get('course_id', type=int)
    selected_course = Course.query.get(course_id) if course_id else (courses[0] if courses else None)
    user_progress = []
    if selected_course:
        for user in selected_course.assigned_users:
            # Video ilerlemesi
            total_videos = len(selected_course.videos)
            completed_videos = 0
            test_score = None
            percent = 0
            if total_videos > 0:
                for video in selected_course.videos:
                    progress = Progress.query.filter_by(user_id=user.id, video_id=video.id).first()
                    if progress and progress.completed:
                        completed_videos += 1
                        if progress.test_score is not None:
                            test_score = progress.test_score
                percent = int((completed_videos / total_videos) * 100)
            user_progress.append({
                'user': user,
                'completed_videos': completed_videos,
                'total_videos': total_videos,
                'percent': percent,
                'test_score': test_score
            })
    return render_template('admin_dashboard.html', courses=courses, selected_course=selected_course, user_progress=user_progress)

@admin_bp.route('/database')
@login_required
@admin_required
def database():
    """Database management"""
    # Database istatistikleri
    stats = {
        'users': User.query.count(),
        'courses': Course.query.count(),
        'videos': Video.query.count(),
        'categories': Category.query.count(),
        'certificate_types': CertificateType.query.count(),
        'certificates': Certificate.query.count(),
        'announcements': Announcement.query.count(),
        'total_progress': Progress.query.count()
    }
    
    # Son eklenen kayıtlar
    recent_users = User.query.order_by(User.created_at.desc()).limit(5).all()
    recent_courses = Course.query.order_by(Course.created_at.desc()).limit(5).all()
    
    return render_template('admin_database.html', stats=stats, recent_users=recent_users, recent_courses=recent_courses)

@admin_bp.route('/database/backup', methods=['POST'])
@login_required
@admin_required
def backup_database():
    """
    Database dosyasını yedekle
    ---
    tags:
      - Admin
    produces:
      - application/octet-stream
    responses:
      302:
        description: Yönlendirme (HTML response)
    """
    try:
        # Mevcut database dosyası
        source_db = 'app.db'  # Yeni projede app.db kullanıyoruz
        if not os.path.exists(source_db):
            flash('Database dosyası bulunamadı!', 'danger')
            return redirect(url_for('admin_database'))
        
        # Yedek dosya adı
        backup_name = f'app_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
        backup_path = os.path.join(current_app.config['UPLOAD_FOLDER'], backup_name)
        
        # Database'i kopyala
        shutil.copy2(source_db, backup_path)
        
        flash(f'Database başarıyla yedeklendi: {backup_name}', 'success')
        
    except Exception as e:
        flash(f'Yedekleme hatası: {str(e)}', 'danger')
    
    return redirect(url_for('admin_database'))

@admin_bp.route('/database/export', methods=['POST'])
@login_required
@admin_required
def export_database():
    """
    Database'i JSON olarak dışa aktır
    ---
    tags:
      - Admin
    produces:
      - application/json
    responses:
      200:
        description: JSON dosyası
        schema:
          type: file
      302:
        description: Yönlendirme (HTML response)
    """
    try:
        # Database bağlantısı
        conn = sqlite3.connect('app.db')
        cursor = conn.cursor()
        
        # Tüm tabloları al
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        
        export_data = {}
        
        for table in tables:
            table_name = table[0]
            cursor.execute(f"SELECT * FROM {table_name}")
            rows = cursor.fetchall()
            
            # Sütun adlarını al
            cursor.execute(f"PRAGMA table_info({table_name});")
            columns = [column[1] for column in cursor.fetchall()]
            
            # Verileri dictionary formatında sakla
            table_data = []
            for row in rows:
                row_dict = {}
                for i, value in enumerate(row):
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    row_dict[columns[i]] = value
                table_data.append(row_dict)
            
            export_data[table_name] = table_data
        
        conn.close()
        
        # JSON dosyası olarak kaydet
        export_filename = f'database_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        export_path = os.path.join(current_app.config['UPLOAD_FOLDER'], export_filename)
        
        with open(export_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
        
        return send_file(export_path, as_attachment=True, download_name=export_filename)
        
    except Exception as e:
        flash(f'Export hatası: {str(e)}', 'danger')
        return redirect(url_for('admin_database'))

@admin_bp.route('/database/import', methods=['POST'])
@login_required
@admin_required
def import_database():
    """
    JSON dosyasından database'i içe aktır
    ---
    tags:
      - Admin
    consumes:
      - multipart/form-data
    parameters:
      - name: database_file
        in: formData
        type: file
        required: true
        description: JSON database dosyası
    responses:
      302:
        description: Yönlendirme (HTML response)
    """
    if 'database_file' not in request.files:
        flash('Dosya seçilmedi!', 'danger')
        return redirect(url_for('admin_database'))
    
    file = request.files['database_file']
    if file.filename == '':
        flash('Dosya seçilmedi!', 'danger')
        return redirect(url_for('admin_database'))
    
    if not file.filename.endswith('.json'):
        flash('Sadece JSON dosyaları kabul edilir!', 'danger')
        return redirect(url_for('admin_database'))
    
    try:
        # JSON dosyasını oku
        data = json.load(file)
        
        # Database bağlantısı
        conn = sqlite3.connect('app.db')
        cursor = conn.cursor()
        
        # Mevcut verileri temizle
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        
        for table in tables:
            table_name = table[0]
            cursor.execute(f"DELETE FROM {table_name}")
        
        # Yeni verileri ekle
        for table_name, table_data in data.items():
            if table_data:
                # İlk satırdan sütun adlarını al
                columns = list(table_data[0].keys())
                placeholders = ', '.join(['?' for _ in columns])
                column_names = ', '.join(columns)
                
                for row in table_data:
                    values = [row.get(col) for col in columns]
                    cursor.execute(f"INSERT INTO {table_name} ({column_names}) VALUES ({placeholders})", values)
        
        conn.commit()
        conn.close()
        
        flash('Database başarıyla import edildi!', 'success')
        
    except Exception as e:
        flash(f'Import hatası: {str(e)}', 'danger')
    
    return redirect(url_for('admin_database'))

@admin_bp.route('/database/reset', methods=['POST'])
@login_required
@admin_required
def reset_database():
    """Reset database to default state"""
    try:
        # Database'i sıfırla
        db.drop_all()
        db.create_all()
        
        # Default admin user oluştur
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                email='admin@adawall.com',
                first_name='Admin',
                last_name='User',
                is_admin=True
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
        
        flash('Database başarıyla sıfırlandı!', 'success')
        
    except Exception as e:
        flash(f'Database sıfırlama hatası: {str(e)}', 'danger')
    
    return redirect(url_for('admin_database'))

@admin_bp.route('/password-requests')
@login_required
@admin_required
def password_requests():
    """Password reset requests"""
    pending_requests = PasswordReset.query.filter_by(status='pending').order_by(PasswordReset.requested_at.desc()).all()
    processed_requests = PasswordReset.query.filter(
        PasswordReset.status.in_(['approved', 'rejected', 'completed'])
    ).order_by(PasswordReset.approved_at.desc()).limit(50).all()
    
    return render_template('admin_password_requests.html', 
                         pending_requests=pending_requests,
                         processed_requests=processed_requests)

@admin_bp.route('/approve-password-request/<int:request_id>', methods=['POST'])
@login_required
@admin_required
def approve_password_request(request_id):
    """Approve password reset request"""
    reset_request = PasswordReset.query.get_or_404(request_id)
    
    if reset_request.status != 'pending':
        flash('Bu talep zaten işleme alınmış.', 'warning')
        return redirect(url_for('admin.password_requests'))
    
    # Generate token for password reset
    token = secrets.token_urlsafe(32)
    
    # Update request
    reset_request.status = 'approved'
    reset_request.approved_by = current_user.id
    reset_request.approved_at = datetime.utcnow()
    reset_request.token = token
    reset_request.expires_at = datetime.utcnow() + timedelta(hours=24)  # 24 saat geçerli
    
    db.session.commit()
    
    flash(f'{reset_request.user.first_name} {reset_request.user.last_name} kullanıcısının şifre sıfırlama talebi onaylandı.', 'success')
    return redirect(url_for('admin.password_requests'))

@admin_bp.route('/reject-password-request/<int:request_id>', methods=['POST'])
@login_required
@admin_required
def reject_password_request(request_id):
    """Reject password reset request"""
    reset_request = PasswordReset.query.get_or_404(request_id)
    
    if reset_request.status != 'pending':
        flash('Bu talep zaten işleme alınmış.', 'warning')
        return redirect(url_for('admin.password_requests'))
    
    # Update request
    reset_request.status = 'rejected'
    reset_request.approved_by = current_user.id
    reset_request.approved_at = datetime.utcnow()
    
    db.session.commit()
    
    flash(f'{reset_request.user.first_name} {reset_request.user.last_name} kullanıcısının şifre sıfırlama talebi reddedildi.', 'info')
    return redirect(url_for('admin.password_requests'))

@admin_bp.route('/download_certificate/<int:certificate_id>')
@login_required
@admin_required
def download_certificate_pdf(certificate_id):
    """
    Admin olarak sertifika PDF dosyasını indir
    ---
    tags:
      - Admin
    parameters:
      - name: certificate_id
        in: path
        type: integer
        required: true
        description: Sertifika ID
    produces:
      - application/pdf
    responses:
      200:
        description: PDF dosyası
        schema:
          type: file
      404:
        description: Dosya bulunamadı
    """
    cert = Certificate.query.get_or_404(certificate_id)
    from flask import send_file, current_app
    import os
    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], cert.certificate_path)
    return send_file(file_path, as_attachment=True) 

@admin_bp.route('/delete_certificate/<int:certificate_id>', methods=['POST'])
@login_required
@admin_required
def delete_certificate(certificate_id):
    cert = Certificate.query.get_or_404(certificate_id)
    db.session.delete(cert)
    db.session.commit()
    flash('Sertifika silindi.', 'success')
    from flask import request, url_for, redirect
    return redirect(request.referrer or url_for('admin_user_certificates')) 