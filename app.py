from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
import os
from datetime import datetime, timedelta
import openpyxl
from config import Config
from models import db, login_manager, User, Course, Video, Test, Question, Option, Progress, Announcement, Category, CertificateType, Certificate, PasswordReset, Pdf, PdfProgress
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import uuid
from flask_talisman import Talisman
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField
from wtforms.validators import DataRequired, Length, Regexp, EqualTo
import secrets
from flask_migrate import Migrate
from flask_mail import Mail, Message
from flask_weasyprint import render_pdf
from flask_weasyprint import HTML
from sqlalchemy.exc import IntegrityError

# Limiter nesnesini yeni sÃ¼rÃ¼me uygun ÅŸekilde baÅŸlat
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"]
)

def create_app():
    app = Flask(__name__)
    app.config.from_object(Config)
    
    # Upload limit'ini yÃ¼ksek set et
    app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB
    
    # Debug iÃ§in config'i logla
    print(f"ğŸ”§ MAX_CONTENT_LENGTH set to: {app.config['MAX_CONTENT_LENGTH'] / (1024*1024)}MB")

    db.init_app(app)
    login_manager.init_app(app)
    login_manager.login_view = 'login'

    # --- GÃ¼venlik: Talisman ve Limiter ---
    Talisman(app, content_security_policy={
        'default-src': ["'self'", 'https://cdn.jsdelivr.net'],
        'img-src': ["'self'", 'data:', 'https://cdn.jsdelivr.net'],
        'script-src': ["'self'", "'unsafe-inline'", 'https://cdn.jsdelivr.net'],
        'style-src': ["'self'", "'unsafe-inline'", 'https://cdn.jsdelivr.net'],
    })
    # Limiter'Ä± burada app ile baÅŸlat
    limiter.init_app(app)

    return app

app = create_app()
mail = Mail(app)
migrate = Migrate(app, db)

# SQLite performance optimization event listener
from sqlalchemy import event
from sqlalchemy.engine import Engine
import sqlite3

@event.listens_for(Engine, "connect")
def set_sqlite_pragma(dbapi_connection, connection_record):
    """SQLite performance optimizations"""
    if isinstance(dbapi_connection, sqlite3.Connection):
        cursor = dbapi_connection.cursor()
        # Performance optimizations
        cursor.execute("PRAGMA journal_mode=WAL")  # Write-Ahead Logging
        cursor.execute("PRAGMA synchronous=NORMAL")  # Faster sync
        cursor.execute("PRAGMA cache_size=-64000")  # 64MB cache
        cursor.execute("PRAGMA temp_store=MEMORY")  # Memory temp storage
        cursor.close()
        print("ğŸ”§ SQLite PRAGMA optimizations applied")

# Error handler for file upload limits
@app.errorhandler(413)
def request_entity_too_large(error):
    """Handle file too large error"""
    max_size_mb = app.config['MAX_CONTENT_LENGTH'] / (1024 * 1024)
    flash(f'âŒ Dosya Ã§ok bÃ¼yÃ¼k! Maximum {max_size_mb:.0f}MB yÃ¼kleyebilirsiniz. DosyalarÄ±nÄ±zÄ± kÃ¼Ã§Ã¼ltÃ¼n.', 'danger')
    return redirect(url_for('new_course'))

# --- GÃ¼Ã§lÃ¼ ÅŸifre politikasÄ± iÃ§in yardÄ±mcÄ± fonksiyon ---
def password_policy_check(password):
    if len(password) < 8:
        return False, "Åifre en az 8 karakter olmalÄ±."
    if not any(c.isupper() for c in password):
        return False, "Åifre en az bir bÃ¼yÃ¼k harf iÃ§ermeli."
    if not any(c.islower() for c in password):
        return False, "Åifre en az bir kÃ¼Ã§Ã¼k harf iÃ§ermeli."
    if not any(c.isdigit() for c in password):
        return False, "Åifre en az bir rakam iÃ§ermeli."
    if not any(c in '!@#$%^&*()_+-=[]{}|;:,.<>?' for c in password):
        return False, "Åifre en az bir Ã¶zel karakter iÃ§ermeli (!@#$%^&* vb.)."
    return True, ""

def send_password_reset_email(user, token):
    """KullanÄ±cÄ±ya ÅŸifre sÄ±fÄ±rlama e-postasÄ± gÃ¶nderir."""
    msg = Message(
        'AdaWall EÄŸitim - Åifre SÄ±fÄ±rlama Ä°steÄŸi',
        sender=app.config['MAIL_DEFAULT_SENDER'],
        recipients=[user.email]
    )
    reset_url = url_for('reset_password_with_token', token=token, _external=True)
    msg.html = render_template('email/reset_password.html', user=user, reset_url=reset_url)
    try:
        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"E-posta gÃ¶nderme hatasÄ±: {e}")
        return False

# --- Form SÄ±nÄ±flarÄ± ---

class EmptyForm(FlaskForm):
    """CSRF korumasÄ± iÃ§in boÅŸ form."""
    pass

class LoginForm(FlaskForm):
    username = StringField('KullanÄ±cÄ± AdÄ±', validators=[DataRequired()])
    password = PasswordField('Åifre', validators=[DataRequired()])

# --- KayÄ±t/Åifre deÄŸiÅŸtir formunda parola politikasÄ± ---
password_policy = [
    Length(min=8, message='Åifre en az 8 karakter olmalÄ±.'),
    Regexp(r'.*[A-Z].*', message='Åifre en az bir bÃ¼yÃ¼k harf iÃ§ermeli.'),
    Regexp(r'.*[a-z].*', message='Åifre en az bir kÃ¼Ã§Ã¼k harf iÃ§ermeli.'),
    Regexp(r'.*[0-9].*', message='Åifre en az bir rakam iÃ§ermeli.'),
    Regexp(r'.*[^A-Za-z0-9].*', message='Åifre en az bir Ã¶zel karakter iÃ§ermeli.')
]

class ResetPasswordForm(FlaskForm):
    password = PasswordField('Yeni Åifre', validators=[DataRequired()] + password_policy)
    confirm = PasswordField('Yeni Åifreyi Onayla', validators=[
        DataRequired(),
        EqualTo('password', message='Åifreler eÅŸleÅŸmelidir.')
    ])

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            flash('KullanÄ±cÄ± adÄ± ve ÅŸifre zorunludur.', 'danger')
            return render_template('login.html')
            
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('KullanÄ±cÄ± adÄ± veya ÅŸifre hatalÄ±!', 'danger')
            return render_template('login.html')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    if current_user.is_admin:
        courses = Course.query.all()
        selected_course_id = request.args.get('course_id', type=int)
        selected_course = Course.query.get(selected_course_id) if selected_course_id else (courses[0] if courses else None)
        user_progress = []
        if selected_course:
            users = selected_course.assigned_users  # SADECE ATANANLAR
            for user in users:
                # Yeni progress sistemi ile hesapla
                progress_details = selected_course.get_user_progress(user)
                
                # Video sayÄ±larÄ±nÄ± ayrÄ±ca hesapla (eski sistem ile uyumluluk iÃ§in)
                video_ids = set(v.id for v in selected_course.videos)
                completed_videos = len({p.video_id for p in user.progress if p.completed and p.video_id in video_ids})
                total_videos = len(video_ids)
                
                user_progress.append({
                    'user': user,
                    'completed_videos': completed_videos,
                    'total_videos': total_videos,
                    'percent': progress_details.progress_percent,  # DÃ¼zeltilmiÅŸ progress yÃ¼zdesi
                    'test_score': progress_details.test_score,
                    'test_completed': progress_details.passed_test,
                })
        return render_template('admin_dashboard.html', courses=courses, selected_course=selected_course, user_progress=user_progress)
    else:
        courses = current_user.assigned_courses
        announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
        return render_template('user_dashboard.html', courses=courses, announcements=announcements)

@app.route('/course/<int:course_id>')
@login_required
def course(course_id):
    # Admin ise kurs yÃ¶netim sayfasÄ±na yÃ¶nlendir
    if current_user.is_admin:
        return redirect(url_for('admin_manage_course', course_id=course_id))
    
    course = Course.query.get_or_404(course_id)
    content = get_course_content(course.id)

    # EÄŸer kursta hiÃ§ iÃ§erik yoksa
    if not content:
        return render_template('course.html', course=course, status='empty', button_text='Kursta Ä°Ã§erik Yok', start_url=url_for('dashboard'))

    # KullanÄ±cÄ±nÄ±n ilerlemesini al
    progress_details = course.get_user_progress(current_user)
    
    # KullanÄ±cÄ±nÄ±n tamamladÄ±ÄŸÄ± videolarÄ± bul
    completed_video_ids = set()
    for progress in current_user.progress:
        if progress.completed and progress.video:
            if progress.video.course_id == course_id:
                completed_video_ids.add(progress.video_id)

    # GÃ¶rÃ¼ntÃ¼lenen PDF'leri al (database'den)
    viewed_pdf_ids = set()
    pdf_progress_list = PdfProgress.query.filter_by(user_id=current_user.id).all()
    for pdf_progress in pdf_progress_list:
        if pdf_progress.pdf.course_id == course_id:
            viewed_pdf_ids.add(pdf_progress.pdf_id)

    # SÄ±radaki iÃ§eriÄŸi bul
    next_item = None
    for item in content:
        if item['type'] == 'video':
            # Video tamamlanmamÄ±ÅŸsa, bu sÄ±radaki adÄ±m
            if item['id'] not in completed_video_ids:
                next_item = item
                break
        elif item['type'] == 'pdf':
            # PDF gÃ¶rÃ¼ntÃ¼lenmemiÅŸse, bu sÄ±radaki adÄ±m
            if item['id'] not in viewed_pdf_ids:
                next_item = item
                break

    # URL ve durumu belirle
    start_url = url_for('dashboard') 
    status = 'not_started'
    button_text = 'EÄŸitime BaÅŸla'
    
    # HiÃ§ baÅŸlanmamÄ±ÅŸ mÄ±?
    has_any_progress = bool(completed_video_ids or viewed_pdf_ids)
    
    if next_item:
        # Sonraki adÄ±m var
        if next_item['type'] == 'video':
            start_url = url_for('video', video_id=next_item['id'])
        else: # pdf
            start_url = url_for('pdf_viewer', pdf_id=next_item['id'])
        
        # BaÅŸlanmÄ±ÅŸ mÄ±?
        if has_any_progress:
            status = 'in_progress'
            button_text = 'EÄŸitime Devam Et'
        else:
            status = 'not_started'
            button_text = 'EÄŸitime BaÅŸla'
    else:
        # Sonraki adÄ±m yok - tÃ¼m iÃ§erik tamamlandÄ± mÄ± kontrol et
        if progress_details.all_content_completed:
            status = 'completed'
            if course.test_required and (course.test_pdf or course.test_images):
                # Test tamamlanmÄ±ÅŸ mÄ± kontrol et
                if not progress_details.passed_test:
                    start_url = url_for('pdf_test', course_id=course.id)
                    button_text = 'Kursu Bitirme Testine Git'
                    status = 'test_required'
                else:
                    button_text = 'Kurs TamamlandÄ±'
                    start_url = url_for('dashboard')
                    status = 'completed'
            else:
                button_text = 'Kurs TamamlandÄ±'
                start_url = url_for('dashboard')
                status = 'completed'
        else:
            # Ä°Ã§erik tamamlanmamÄ±ÅŸ ama sonraki adÄ±m bulunamadÄ± - bu bir hata durumu
            # En son eriÅŸilebilir iÃ§eriÄŸe yÃ¶nlendir
            if has_any_progress:
                # Son tamamlanan iÃ§eriÄŸi bul
                last_accessible_item = None
                for item in content:
                    if item['type'] == 'video' and item['id'] in completed_video_ids:
                        last_accessible_item = item
                    elif item['type'] == 'pdf' and item['id'] in viewed_pdf_ids:
                        last_accessible_item = item
                
                if last_accessible_item:
                    if last_accessible_item['type'] == 'video':
                        start_url = url_for('video', video_id=last_accessible_item['id'])
                    else:
                        start_url = url_for('pdf_viewer', pdf_id=last_accessible_item['id'])
                    status = 'in_progress'
                    button_text = 'EÄŸitime Devam Et'
                else:
                    # Ä°lk iÃ§eriÄŸe yÃ¶nlendir
                    first_item = content[0]
                    if first_item['type'] == 'video':
                        start_url = url_for('video', video_id=first_item['id'])
                    else:
                        start_url = url_for('pdf_viewer', pdf_id=first_item['id'])
                    status = 'in_progress'
                    button_text = 'EÄŸitime Devam Et'
            else:
                # HiÃ§ baÅŸlanmamÄ±ÅŸ, ilk iÃ§eriÄŸe yÃ¶nlendir
                first_item = content[0]
                if first_item['type'] == 'video':
                    start_url = url_for('video', video_id=first_item['id'])
                else:
                    start_url = url_for('pdf_viewer', pdf_id=first_item['id'])
                status = 'not_started'
                button_text = 'EÄŸitime BaÅŸla'

    return render_template('course.html', course=course, start_url=start_url, status=status, button_text=button_text)

def get_course_content(course_id):
    course = Course.query.get_or_404(course_id)
    videos = [{'type': 'video', 'id': v.id, 'data': v} for v in course.videos]
    pdfs = [{'type': 'pdf', 'id': p.id, 'data': p} for p in course.pdfs]
    return sorted(videos + pdfs, key=lambda x: x['data'].order)

@app.route('/video/<int:video_id>')
@login_required
def video(video_id):
    video = Video.query.get_or_404(video_id)
    course = video.course
    
    content_list = get_course_content(course.id)
    current_index = -1
    for i, item in enumerate(content_list):
        if item['type'] == 'video' and item['id'] == video_id:
            current_index = i
            break
    
    # KullanÄ±cÄ±nÄ±n tamamladÄ±ÄŸÄ± videolarÄ± bul
    completed_video_ids = set()
    for progress in current_user.progress:
        if progress.completed and progress.video:
            if progress.video.course_id == course.id:
                completed_video_ids.add(progress.video_id)
    
    # GÃ¶rÃ¼ntÃ¼lenen PDF'leri al (database'den)
    viewed_pdf_ids = set()
    pdf_progress_list = PdfProgress.query.filter_by(user_id=current_user.id).all()
    for pdf_progress in pdf_progress_list:
        if pdf_progress.pdf.course_id == course.id:
            viewed_pdf_ids.add(pdf_progress.pdf_id)
    
    # Ã–nceki ve sonraki URL'leri belirle
    prev_url = None
    next_url = None
    can_access_next = False
    
    # Ã–nceki iÃ§erik (her zaman eriÅŸilebilir)
    if current_index > 0:
        prev_item = content_list[current_index - 1]
        if prev_item['type'] == 'video':
            prev_url = url_for('video', video_id=prev_item['id'])
        else:
            prev_url = url_for('pdf_viewer', pdf_id=prev_item['id'])
    
    # Sonraki iÃ§erik belirleme
    current_video_completed = video_id in completed_video_ids
    
    if current_index < len(content_list) - 1:
        # HenÃ¼z son iÃ§erik deÄŸil, sonraki video/PDF var
        if current_video_completed:
            next_item = content_list[current_index + 1]
            can_access_next = True
            if next_item['type'] == 'video':
                next_url = url_for('video', video_id=next_item['id'])
            else:
                next_url = url_for('pdf_viewer', pdf_id=next_item['id'])
    else:
        # Bu son iÃ§erik, test kontrolÃ¼ yap
        if current_video_completed:
            # Bu video tamamlandÄ±ysa, test var mÄ± kontrol et
            if course.test_required and (course.test_pdf or course.test_images):
                # Test tamamlanmÄ±ÅŸ mÄ± kontrol et
                existing_progress = Progress.query.filter_by(user_id=current_user.id, video_id=video_id).first()
                if not existing_progress or not existing_progress.test_completed:
                    next_url = url_for('pdf_test', course_id=course.id)
                    can_access_next = True

    progress_details = course.get_user_progress(current_user)
    video_progress = video.get_progress(current_user)

    return render_template(
        'video.html',
        video=video,
        course=course,
        progress=video_progress,
        completed_steps=progress_details.completed_steps,
        total_steps=progress_details.total_steps,
        progress_percent=progress_details.progress_percent,
        prev_url=prev_url,
        next_url=next_url,
        can_access_next=can_access_next
    )

@app.route('/pdf/<int:pdf_id>')
@login_required
def pdf_viewer(pdf_id):
    pdf = Pdf.query.get_or_404(pdf_id)
    course = pdf.course

    # PDF gÃ¶rÃ¼ntÃ¼leme kaydÄ±nÄ± oluÅŸtur/gÃ¼ncelle (database'de)
    existing_progress = PdfProgress.query.filter_by(user_id=current_user.id, pdf_id=pdf_id).first()
    if not existing_progress:
        pdf_progress = PdfProgress(user_id=current_user.id, pdf_id=pdf_id)
        db.session.add(pdf_progress)
        db.session.commit()

    content_list = get_course_content(course.id)
    current_index = -1
    for i, item in enumerate(content_list):
        if item['type'] == 'pdf' and item['id'] == pdf_id:
            current_index = i
            break
    
    # Navigasyon URL'leri
    prev_url = None
    next_url = None
    
    # Ã–nceki iÃ§erik
    if current_index > 0:
        prev_item = content_list[current_index - 1]
        if prev_item['type'] == 'video':
            prev_url = url_for('video', video_id=prev_item['id'])
        else:
            prev_url = url_for('pdf_viewer', pdf_id=prev_item['id'])
    
    # Sonraki iÃ§erik belirleme
    if current_index < len(content_list) - 1:
        # HenÃ¼z son iÃ§erik deÄŸil, sonraki video/PDF var
        next_item = content_list[current_index + 1]
        if next_item['type'] == 'video':
            next_url = url_for('video', video_id=next_item['id'])
        else:
            next_url = url_for('pdf_viewer', pdf_id=next_item['id'])
    else:
        # Bu son iÃ§erik, test kontrolÃ¼ yap
        progress_details = course.get_user_progress(current_user)
        if course.test_required and progress_details.all_content_completed:
            if (course.test_pdf or course.test_images):
                # Test tamamlanmÄ±ÅŸ mÄ± kontrol et
                course_videos = sorted(course.videos, key=lambda v: v.order)
                last_video = course_videos[-1] if course_videos else None
                test_progress = None
                if last_video:
                    test_progress = Progress.query.filter_by(user_id=current_user.id, video_id=last_video.id).first()
                
                if not test_progress or not test_progress.test_completed:
                    next_url = url_for('pdf_test', course_id=course.id)

    # Progress bilgilerini al
    progress_details = course.get_user_progress(current_user)

    return render_template('pdf_viewer.html', 
                         pdf=pdf, 
                         course=course, 
                         prev_url=prev_url, 
                         next_url=next_url,
                         completed_steps=progress_details.completed_steps,
                         total_steps=progress_details.total_steps,
                         progress_percent=progress_details.progress_percent)

@app.route('/video/<int:video_id>/complete', methods=['POST'])
@login_required
def complete_video_and_get_next(video_id):
    """Videoyu tamamlandÄ± olarak iÅŸaretler ve bir sonraki adÄ±mÄ±n URL'sini dÃ¶ndÃ¼rÃ¼r."""
    video = Video.query.get_or_404(video_id)
    
    # Mevcut progress kaydÄ±nÄ± kontrol et
    progress = Progress.query.filter_by(user_id=current_user.id, video_id=video_id).first()
    
    if not progress:
        # Yeni progress kaydÄ± oluÅŸtur
        progress = Progress(user_id=current_user.id, video_id=video_id)
        db.session.add(progress)
    
    # Progress'i gÃ¼ncelle
    progress.completed = True
    progress.completed_at = datetime.utcnow()
    
    try:
        db.session.commit()
    except IntegrityError:
        # EÄŸer UNIQUE constraint hatasÄ± alÄ±rsak, rollback yap ve mevcut kaydÄ± gÃ¼ncelle
        db.session.rollback()
        # Mevcut kaydÄ± tekrar al ve gÃ¼ncelle
        progress = Progress.query.filter_by(user_id=current_user.id, video_id=video_id).first()
        if progress:
            progress.completed = True
            progress.completed_at = datetime.utcnow()
            db.session.commit()

    course = video.course
    content_list = get_course_content(course.id)
    
    # Mevcut video'nun content listesindeki pozisyonunu bul
    current_index = -1
    for i, item in enumerate(content_list):
        if item['type'] == 'video' and item['id'] == video_id:
            current_index = i
            break
    
    next_url = None
    message = "Video tamamlandÄ±!"

    # Sonraki iÃ§erik var mÄ±?
    if current_index != -1 and current_index < len(content_list) - 1:
        next_item = content_list[current_index + 1]
        if next_item['type'] == 'video':
            next_url = url_for('video', video_id=next_item['id'])
            message = "Sonraki videoya geÃ§ebilirsiniz."
        else:  # PDF
            next_url = url_for('pdf_viewer', pdf_id=next_item['id'])
            message = "Sonraki PDF materyaline geÃ§ebilirsiniz."
    else:
        # TÃ¼m iÃ§erik tamamlandÄ±, test var mÄ± kontrol et
        user_progress_details = course.get_user_progress(current_user)
        if course.test_required and user_progress_details.all_content_completed:
            if (course.test_pdf or course.test_images):
                # Test henÃ¼z tamamlanmamÄ±ÅŸsa teste yÃ¶nlendir
                course_videos = sorted(course.videos, key=lambda v: v.order)
                last_video = course_videos[-1] if course_videos else None
                test_progress = None
                if last_video:
                    test_progress = Progress.query.filter_by(user_id=current_user.id, video_id=last_video.id).first()
                
                if not test_progress or not test_progress.test_completed:
                    next_url = url_for('pdf_test', course_id=course.id)
                    message = "TÃ¼m iÃ§erik tamamlandÄ±! Teste geÃ§ebilirsiniz."
                else:
                    message = "Tebrikler, kursu tamamladÄ±nÄ±z!"
            else:
                message = "Tebrikler, kursu tamamladÄ±nÄ±z!"
        else:
            message = "Tebrikler, kursu tamamladÄ±nÄ±z!"
            
    return jsonify({
        'success': True, 
        'next_url': next_url,
        'message': message
    })

@app.route('/test/<int:test_id>', methods=['GET', 'POST'])
@login_required
def test(test_id):
    test = Test.query.get_or_404(test_id)
    if request.method == 'POST':
        score = 0
        for question in test.questions:
            answer = request.form.get(f'question_{question.id}')
            if answer and Option.query.get(int(answer)).is_correct:
                score += 1
        
        progress = Progress.query.filter_by(user_id=current_user.id, video_id=test.course.videos[-1].id).first()
        if progress:
            progress.test_score = score
            progress.test_completed = score >= test.passing_score
            progress.completed_at = datetime.utcnow()
            db.session.commit()
        
        return redirect(url_for('course', course_id=test.course_id))
    
    return render_template('test.html', test=test)

@app.route('/admin/new-course', methods=['GET', 'POST'])
@login_required
def new_course():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        # Quick file check - performance optimized
        print(f"ğŸš€ Starting course creation process...")
        # Temel kurs bilgilerini al
        title = request.form.get('title')
        description = request.form.get('description')
        category_id = request.form.get('category_id')
        assigned_user_ids = request.form.getlist('assigned_users')
        passing_score = request.form.get('passing_score', 70, type=int)
        test_required = 'test_required' in request.form

        # Yeni kursu oluÅŸtur ve session'a ekle
        new_course = Course(
            title=title, 
            description=description, 
            category_id=category_id,
            passing_score=passing_score,
            test_required=test_required,
            certificate_type_id=None # Åimdilik boÅŸ, gerekirse gÃ¼ncellenebilir
        )
        db.session.add(new_course)
        
        # Performance Critical: Single Transaction Approach
        try:
            # KullanÄ±cÄ±larÄ± ata
            new_course.assigned_users = User.query.filter(User.id.in_(assigned_user_ids)).all()

            # Ä°lk commit - kurs ID'sini al
            db.session.commit()
            print(f"âœ… Course created with ID: {new_course.id}")

            # Dosya iÅŸlemleri - optimized approach
            content_titles = request.form.getlist('content_titles[]')
            content_types = request.form.getlist('content_types[]')
            content_orders = request.form.getlist('content_orders[]')
            content_files = request.files.getlist('content_files[]')

            # Memory efficient processing
            items_to_add = []
            upload_folder = app.config['UPLOAD_FOLDER']
            
            print(f"ğŸ“ Processing {len(content_files)} files...")
            
            for i, file in enumerate(content_files):
                if file and file.filename:
                    content_type = content_types[i]
                    timestamp = int(datetime.utcnow().timestamp())
                    
                    # Fast filename generation
                    filename = f"{content_type}_{new_course.id}_{timestamp}_{secure_filename(file.filename)}"
                    file_path = os.path.join(upload_folder, filename)
                    
                    # Direct file save - no size checking for speed
                    try:
                        print(f"â¬†ï¸  Saving {filename}...")
                        file.save(file_path)
                        
                        # Create database object
                        if content_type == 'video':
                            item = Video(
                                title=content_titles[i],
                                video_path=filename,
                                course_id=new_course.id,
                                order=int(content_orders[i])
                            )
                        else:  # pdf
                            item = Pdf(
                                title=content_titles[i],
                                pdf_path=filename,
                                course_id=new_course.id,
                                order=int(content_orders[i])
                            )
                        
                        items_to_add.append(item)
                        print(f"âœ… {filename} processed")
                        
                    except Exception as e:
                        print(f"âŒ Error saving {filename}: {e}")
                        flash(f'Dosya kaydedilemedi: {file.filename}', 'warning')
            
            # Single batch insert for all items
            if items_to_add:
                print(f"ğŸ’¾ Batch inserting {len(items_to_add)} items to database...")
                db.session.add_all(items_to_add)
            
            # Test dosyasÄ± iÅŸleme - Speed optimized
            if new_course.test_required:
                test_file = request.files.get('test_file')
                pdf_question_count = request.form.get('pdf_question_count')
                pdf_answer_key = request.form.get('pdf_answer_key')

                if test_file and test_file.filename and pdf_question_count and pdf_answer_key:
                    ext = os.path.splitext(test_file.filename)[1].lower()
                    try:
                        # Fast test file processing
                        if ext == '.pdf':
                            filename = f"test_{new_course.id}_{secure_filename(test_file.filename)}"
                            new_course.test_pdf = filename
                            new_course.test_images = None
                        elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
                            filename = f"testimg_{new_course.id}_{int(datetime.utcnow().timestamp())}_{secure_filename(test_file.filename)}"
                            new_course.test_images = filename
                            new_course.test_pdf = None
                        else:
                            filename = None
                        
                        if filename:
                            print(f"â¬†ï¸  Saving test file: {filename}")
                            test_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                            new_course.test_question_count = int(pdf_question_count)
                            new_course.test_answer_key = pdf_answer_key
                            print(f"âœ… Test file saved")
                    except Exception as e:
                        print(f"âŒ Test file error: {e}")
                        flash(f'Test dosyasÄ± kaydedilemedi: {str(e)}', 'warning')
            
            # Final commit - everything in one transaction
            print(f"ğŸ’¾ Final commit...")
            db.session.commit()
            print(f"ğŸ‰ Course creation completed successfully!")
            flash('ğŸ‰ Kurs baÅŸarÄ±yla oluÅŸturuldu!', 'success')
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            print(f"âŒ Course creation error: {e}")
            db.session.rollback()
            flash(f'Kurs oluÅŸturma hatasÄ±: {str(e)}', 'danger')
            return redirect(url_for('new_course'))

    # GET request iÃ§in
    users = User.query.filter_by(is_admin=False).all()
    categories = Category.query.all()
    return render_template('new_course.html', users=users, categories=categories)

@app.route('/admin/course/<int:course_id>/upload', methods=['POST'])
@login_required
def upload_video(course_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    if 'video' not in request.files:
        flash('No video file')
        return redirect(url_for('course', course_id=course_id))
    
    file = request.files['video']
    if file.filename == '':
        flash('No selected file')
        return redirect(url_for('course', course_id=course_id))
    
    if file:
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        
        video = Video(
            title=request.form.get('title'),
            video_path=filename,
            course_id=course_id,
            order=Video.query.filter_by(course_id=course_id).count() + 1
        )
        db.session.add(video)
        db.session.commit()
    
    return redirect(url_for('course', course_id=course_id))

@app.route('/admin/course/<int:course_id>/pdf/upload', methods=['POST'])
@login_required
def upload_pdf(course_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    course = Course.query.get_or_404(course_id)
    if 'pdf' not in request.files:
        flash('PDF dosyasÄ± bulunamadÄ±.', 'danger')
        return redirect(url_for('edit_course', course_id=course_id))
    
    file = request.files['pdf']
    if file.filename == '':
        flash('Dosya seÃ§ilmedi.', 'danger')
        return redirect(url_for('edit_course', course_id=course_id))

    if file:
        filename = secure_filename(f"pdf_{course_id}_{datetime.utcnow().timestamp()}_{file.filename}")
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        
        new_pdf = Pdf(
            title=request.form.get('title'),
            pdf_path=filename,
            course_id=course_id,
            order=int(request.form.get('order'))
        )
        db.session.add(new_pdf)
        db.session.commit()
        flash('PDF materyal baÅŸarÄ±yla eklendi.', 'success')
    
    return redirect(url_for('edit_course', course_id=course_id))

@app.route('/admin/report')
@login_required
def generate_report():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))

    selected_course_id = request.args.get('course_id', type=int)
    selected_course = Course.query.get(selected_course_id) if selected_course_id else None

    if not selected_course:
        flash('Rapor iÃ§in bir kurs seÃ§melisiniz.')
        return redirect(url_for('dashboard'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EÄŸitim Raporu"

    # BaÅŸlÄ±klarÄ± ekle
    headers = ['KullanÄ±cÄ±', 'E-posta', 'Kurs', 'Kategori', 'Tamamlanan Ä°Ã§erik', 'Toplam Ä°Ã§erik', 'Ä°lerleme (%)', 'Test Sonucu', 'Durum', 'Tamamlanma Tarihi']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='6A82FB', end_color='6A82FB', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin', color='B4B4B4'), right=Side(style='thin', color='B4B4B4'), top=Side(style='thin', color='B4B4B4'), bottom=Side(style='thin', color='B4B4B4'))

    row = 2
    users = selected_course.assigned_users
    for user in users:
        # Yeni progress sistemi ile hesapla
        progress_details = selected_course.get_user_progress(user)
        
        # Test sonucu ve durumu
        test_score = '-'
        test_status = 'Devam Ediyor'
        last_completed_at = ''
        
        if progress_details.passed_test:
            test_score = progress_details.test_score if progress_details.test_score is not None else '-'
            test_status = 'TamamlandÄ±'
        elif progress_details.is_completed:
            test_status = 'TamamlandÄ±'
        
        # Son tamamlanma tarihini bul
        latest_date = None
        for video in selected_course.videos:
            progress = Progress.query.filter_by(user_id=user.id, video_id=video.id).first()
            if progress and progress.completed and progress.completed_at:
                if latest_date is None or progress.completed_at > latest_date:
                    latest_date = progress.completed_at
        
        if latest_date:
            last_completed_at = latest_date.strftime('%Y-%m-%d %H:%M:%S')
        
        ws.cell(row=row, column=1, value=f"{user.first_name} {user.last_name}")
        ws.cell(row=row, column=2, value=user.email)
        ws.cell(row=row, column=3, value=selected_course.title)
        ws.cell(row=row, column=4, value=selected_course.category.name if selected_course.category else '')
        ws.cell(row=row, column=5, value=progress_details.completed_steps)
        ws.cell(row=row, column=6, value=progress_details.total_steps)
        percent_cell = ws.cell(row=row, column=7, value=f"{progress_details.progress_percent}%")
        percent_cell.alignment = Alignment(horizontal='center')
        percent_cell.fill = PatternFill(start_color='A1C4FD', end_color='A1C4FD', fill_type='solid')
        ws.cell(row=row, column=8, value=test_score)
        status_cell = ws.cell(row=row, column=9, value=test_status)
        if test_status == 'TamamlandÄ±':
            status_cell.fill = PatternFill(start_color='43E97B', end_color='43E97B', fill_type='solid')
            status_cell.font = Font(bold=True, color='222222')
        else:
            status_cell.fill = PatternFill(start_color='FFB347', end_color='FFB347', fill_type='solid')
            status_cell.font = Font(bold=True, color='222222')
        ws.cell(row=row, column=10, value=last_completed_at)
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = Border(left=Side(style='thin', color='B4B4B4'), right=Side(style='thin', color='B4B4B4'), top=Side(style='thin', color='B4B4B4'), bottom=Side(style='thin', color='B4B4B4'))
        row += 1

    # SÃ¼tun geniÅŸliklerini ayarla
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 4
    report_path = os.path.join(app.config['UPLOAD_FOLDER'], f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    wb.save(report_path)
    return send_file(report_path, as_attachment=True)

def check_course_completion(user, course):
    """
    KullanÄ±cÄ±nÄ±n belirli bir kursu tamamlayÄ±p tamamlamadÄ±ÄŸÄ±nÄ± kontrol eder.
    Bu mantÄ±k artÄ±k Course modelindeki 'get_user_progress' metoduna taÅŸÄ±ndÄ±.
    """
    return course.get_user_progress(user).is_completed

@app.route('/complete-video/<int:video_id>', methods=['POST'])
@login_required
def complete_video(video_id):
    """Video tamamlama isteÄŸini (AJAX/Fetch) iÅŸler ve JSON dÃ¶ner."""
    video = Video.query.get_or_404(video_id)
    progress = Progress.query.filter_by(user_id=current_user.id, video_id=video_id).first()
    
    if not progress:
        progress = Progress(user_id=current_user.id, video_id=video_id)
        db.session.add(progress)

    progress.completed = True
    progress.completed_at = datetime.utcnow()
    
    try:
        db.session.commit()
    except IntegrityError:
        # EÄŸer UNIQUE constraint hatasÄ± alÄ±rsak, rollback yap ve mevcut kaydÄ± gÃ¼ncelle
        db.session.rollback()
        # Mevcut kaydÄ± tekrar al ve gÃ¼ncelle
        progress = Progress.query.filter_by(user_id=current_user.id, video_id=video_id).first()
        if progress:
            progress.completed = True
            progress.completed_at = datetime.utcnow()
            db.session.commit()

    return jsonify({'success': True, 'message': 'Ä°lerleme kaydedildi.'})

@app.route('/api/course/<int:course_id>/progress')
@login_required
def api_course_progress(course_id):
    """Kurs ilerlemesini kontrol eden API endpoint'i."""
    course = Course.query.get_or_404(course_id)
    
    # KullanÄ±cÄ±nÄ±n bu kurstaki video ilerlemelerini al
    video_progress = {}
    completed_videos = 0
    
    for video in course.videos:
        progress = Progress.query.filter_by(
            user_id=current_user.id, 
            video_id=video.id
        ).first()
        
        video_progress[video.id] = progress.completed if progress else False
        if video_progress[video.id]:
            completed_videos += 1
    
    # Kurs bilgilerini hazÄ±rla
    course_data = {
        'id': course.id,
        'title': course.title,
        'test_pdf': course.test_pdf,
        'test_images': course.test_images,
        'videos': [{'id': v.id, 'title': v.title, 'order': v.order} for v in course.videos]
    }
    
    return jsonify({
        'success': True,
        'course': course_data,
        'completed_videos': completed_videos,
        'total_videos': len(course.videos),
        'video_progress': video_progress
    })

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm = request.form.get('confirm')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        is_admin = request.form.get('is_admin') == 'on'
        admin_password = request.form.get('admin_password')

        if not username or not email or not password or not confirm or not first_name or not last_name:
            flash('TÃ¼m alanlarÄ± doldurun.', 'danger')
            return render_template('register.html')
        if password != confirm:
            flash('Åifreler eÅŸleÅŸmiyor.', 'danger')
            return render_template('register.html')
            
        # GÃ¼Ã§lÃ¼ ÅŸifre kontrolÃ¼
        is_valid, message = password_policy_check(password)
        if not is_valid:
            flash(message, 'danger')
            return render_template('register.html')
            
        if User.query.filter_by(username=username).first():
            flash('Bu kullanÄ±cÄ± adÄ± zaten alÄ±nmÄ±ÅŸ.', 'danger')
            return render_template('register.html')
        if User.query.filter_by(email=email).first():
            flash('Bu e-posta zaten kayÄ±tlÄ±.', 'danger')
            return render_template('register.html')
        
        # Admin ÅŸifre kontrolÃ¼ - SADECE admin olarak kayÄ±t olunuyorsa
        if is_admin:
            if not admin_password or admin_password != app.config['ADMIN_REGISTRATION_KEY']:
                flash('GeÃ§ersiz admin kayÄ±t ÅŸifresi.', 'danger')
                return render_template('register.html')

        user = User(username=username, email=email, first_name=first_name, last_name=last_name, is_admin=is_admin)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash('KayÄ±t baÅŸarÄ±lÄ±! GiriÅŸ yapabilirsiniz.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/delete-account', methods=['POST'])
@login_required
def delete_account():
    # KullanÄ±cÄ±nÄ±n tÃ¼m ilerleme kayÄ±tlarÄ±nÄ± sil
    Progress.query.filter_by(user_id=current_user.id).delete()
    
    # KullanÄ±cÄ±yÄ± sil
    db.session.delete(current_user)
    db.session.commit()
    
    logout_user()
    flash('HesabÄ±nÄ±z baÅŸarÄ±yla silindi.')
    return redirect(url_for('index'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        # Profil gÃ¼ncelleme
        if 'update_profile' in request.form:
            email = request.form.get('email')
            if email != current_user.email and User.query.filter_by(email=email).first():
                flash('Bu e-posta adresi zaten kullanÄ±lÄ±yor.')
                return redirect(url_for('profile'))
            
            current_user.email = email
            db.session.commit()
            flash('Profil bilgileriniz gÃ¼ncellendi.')
            return redirect(url_for('profile'))
        
        # Åifre deÄŸiÅŸtirme
        elif 'change_password' in request.form:
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            if not current_user.check_password(current_password):
                flash('Mevcut ÅŸifreniz yanlÄ±ÅŸ.')
                return redirect(url_for('profile'))
            
            if new_password != confirm_password:
                flash('Yeni ÅŸifreler eÅŸleÅŸmiyor.')
                return redirect(url_for('profile'))
            
            current_user.set_password(new_password)
            db.session.commit()
            flash('Åifreniz baÅŸarÄ±yla deÄŸiÅŸtirildi.')
            return redirect(url_for('profile'))
    
    return render_template('profile.html')

@app.route('/admin/users')
@login_required
def admin_users():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    users = User.query.all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/user/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        user.first_name = request.form.get('first_name')
        user.last_name = request.form.get('last_name')
        user.email = request.form.get('email')
        user.is_admin = True if request.form.get('is_admin') == 'on' else False
        db.session.commit()
        flash('KullanÄ±cÄ± bilgileri gÃ¼ncellendi.', 'success')
        return redirect(url_for('admin_users'))
    return render_template('edit_user.html', user=user)

@app.route('/admin/user/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    user = User.query.get_or_404(user_id)
    if user.is_admin:
        flash('Admin kullanÄ±cÄ±lar silinemez.', 'danger')
        return redirect(url_for('admin_users'))
    # KullanÄ±cÄ±ya ait tÃ¼m ilerlemeleri sil
    Progress.query.filter_by(user_id=user.id).delete()
    db.session.delete(user)
    db.session.commit()
    flash('KullanÄ±cÄ± baÅŸarÄ±yla silindi.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/edit-course/<int:course_id>', methods=['GET', 'POST'])
@login_required
def edit_course(course_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    course = Course.query.get_or_404(course_id)
    if request.method == 'POST':
        course.title = request.form.get('title')
        course.description = request.form.get('description')
        course.category_id = request.form.get('category_id')
        course.certificate_type_id = request.form.get('certificate_type_id') or None
        course.passing_score = request.form.get('passing_score', 70, type=int)
        course.test_required = 'test_required' in request.form

        if course.certificate_type_id == 'yok':
            course.certificate_type_id = None

        # KullanÄ±cÄ± atama
        assigned_user_ids = request.form.getlist('assigned_users')
        course.assigned_users = User.query.filter(User.id.in_(assigned_user_ids)).all()
        
        db.session.commit()
        flash('Kurs baÅŸarÄ±yla gÃ¼ncellendi.', 'success')
        return redirect(url_for('admin_manage_course', course_id=course_id))
    
    users = User.query.filter_by(is_admin=False).all()
    categories = Category.query.all()
    certificate_types = CertificateType.query.all()
    return render_template('edit_course.html', course=course, users=users, categories=categories, certificate_types=certificate_types)

@app.route('/admin/course/<int:course_id>/delete', methods=['POST'])
@login_required
def delete_course(course_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    course = Course.query.get_or_404(course_id)
    upload_folder = app.config['UPLOAD_FOLDER']
    
    # Ã–nce testleri, sorularÄ± ve ÅŸÄ±klarÄ± sil
    for test in course.tests:
        for question in test.questions:
            for option in question.options:
                db.session.delete(option)
            db.session.delete(question)
        db.session.delete(test)
    
    # Kursun tÃ¼m videolarÄ±nÄ± ve ilerlemeleri sil
    for video in course.videos:
        # Video dosyasÄ±nÄ± sil
        video_path = os.path.join(upload_folder, video.video_path)
        if os.path.exists(video_path):
            try:
                os.remove(video_path)
                print(f"Video dosyasÄ± silindi: {video_path}")
            except Exception as e:
                print(f"Video dosyasÄ± silinemedi: {video_path} - {e}")
        Progress.query.filter_by(video_id=video.id).delete()
        db.session.delete(video)
    
    # Kursun tÃ¼m PDF dosyalarÄ±nÄ± sil
    for pdf in course.pdfs:
        # PDF progress kayÄ±tlarÄ±nÄ± sil
        PdfProgress.query.filter_by(pdf_id=pdf.id).delete()
        
        # PDF dosyasÄ±nÄ± sil
        pdf_path = os.path.join(upload_folder, pdf.pdf_path)
        if os.path.exists(pdf_path):
            try:
                os.remove(pdf_path)
                print(f"PDF dosyasÄ± silindi: {pdf_path}")
            except Exception as e:
                print(f"PDF dosyasÄ± silinemedi: {pdf_path} - {e}")
        db.session.delete(pdf)
    
    # Test PDF dosyasÄ±nÄ± sil
    if course.test_pdf:
        test_pdf_path = os.path.join(upload_folder, course.test_pdf)
        if os.path.exists(test_pdf_path):
            try:
                os.remove(test_pdf_path)
                print(f"Test PDF dosyasÄ± silindi: {test_pdf_path}")
            except Exception as e:
                print(f"Test PDF dosyasÄ± silinemedi: {test_pdf_path} - {e}")
    
    # Test resim dosyasÄ±nÄ± sil (tek dosya)
    if course.test_images:
        test_img_path = os.path.join(upload_folder, course.test_images)
        if os.path.exists(test_img_path):
            try:
                os.remove(test_img_path)
                print(f"Test resim dosyasÄ± silindi: {test_img_path}")
            except Exception as e:
                print(f"Test resim dosyasÄ± silinemedi: {test_img_path} - {e}")
    
    # Rapor dosyalarÄ±nÄ± sil
    # Tekli rapor
    report_path = os.path.join(upload_folder, f'report_{course.id}.xlsx')
    if os.path.exists(report_path):
        try:
            os.remove(report_path)
            print(f"Rapor dosyasÄ± silindi: {report_path}")
        except Exception as e:
            print(f"Rapor dosyasÄ± silinemedi: {report_path} - {e}")
    
    # Ã‡oklu raporlar (multi_report_...xlsx) - tÃ¼m raporlarÄ± sil
    try:
        for fname in os.listdir(upload_folder):
            if fname.startswith('multi_report_') and fname.endswith('.xlsx'):
                multi_report_path = os.path.join(upload_folder, fname)
                try:
                    os.remove(multi_report_path)
                    print(f"Multi rapor dosyasÄ± silindi: {multi_report_path}")
                except Exception as e:
                    print(f"Multi rapor dosyasÄ± silinemedi: {multi_report_path} - {e}")
    except Exception as e:
        print(f"Upload klasÃ¶rÃ¼ okunamadÄ±: {e}")
    
    # Database'den kursu sil
    db.session.delete(course)
    db.session.commit()
    flash('Kurs ve tÃ¼m dosyalarÄ± baÅŸarÄ±yla silindi.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/admin/courses')
@login_required
def admin_courses():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    courses = Course.query.all()
    return render_template('admin_courses.html', courses=courses)

@app.route('/admin/course/<int:course_id>/test/edit', methods=['POST'])
@login_required
def edit_test(course_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    course = Course.query.get_or_404(course_id)
    title = request.form.get('title')
    passing_score = int(request.form.get('passing_score'))

    # Test var mÄ± kontrol et
    test = Test.query.filter_by(course_id=course_id).first()
    if not test:
        test = Test(title=title, course_id=course_id, passing_score=passing_score)
        db.session.add(test)
        db.session.commit()
    else:
        test.title = title
        test.passing_score = passing_score
        # Eski sorularÄ± ve ÅŸÄ±klarÄ± sil
        for q in test.questions:
            for o in q.options:
                db.session.delete(o)
            db.session.delete(q)
        db.session.commit()

    # SorularÄ± ve ÅŸÄ±klarÄ± ekle
    questions = []
    i = 0
    while True:
        q_text = request.form.get(f'question_{i}')
        if not q_text:
            break
        question = Question(text=q_text, test_id=test.id)
        db.session.add(question)
        db.session.commit()
        # ÅÄ±klar
        options = []
        correct_index = request.form.get(f'correct_{i}')
        for j in range(2):
            o_text = request.form.get(f'option_{i}_{j}')
            if o_text:
                is_correct = (str(j) == correct_index)
                option = Option(text=o_text, is_correct=is_correct, question_id=question.id)
                db.session.add(option)
        db.session.commit()
        i += 1

    flash('Test baÅŸarÄ±yla kaydedildi.')
    return redirect(url_for('course', course_id=course_id))

@app.route('/admin/course/<int:course_id>/testpdf', methods=['POST'])
@login_required
def upload_test_pdf(course_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    course = Course.query.get_or_404(course_id)
    file = request.files.get('test_pdf')
    question_count = int(request.form.get('question_count'))
    answer_key = request.form.get('answer_key').replace(' ', '').upper()  # e.g. A,B,C,D

    # PDF dosyasÄ±nÄ± kaydet
    if file and file.filename.endswith('.pdf'):
        filename = f"test_{course_id}_{secure_filename(file.filename)}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        # Test PDF ve anahtarÄ±nÄ± kursa kaydet (basit Ã§Ã¶zÃ¼m: Course modeline alan ekle)
        course.test_pdf = filename
        course.test_question_count = question_count
        course.test_answer_key = answer_key
        db.session.commit()
        flash('Test PDF ve cevap anahtarÄ± kaydedildi.')
    else:
        flash('GeÃ§erli bir PDF dosyasÄ± seÃ§melisiniz.')
    return redirect(url_for('course', course_id=course_id))

@app.route('/admin/certificate/<int:certificate_id>/upload', methods=['POST'])
@login_required
def admin_upload_certificate_file(certificate_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    certificate = Certificate.query.get_or_404(certificate_id)
    file = request.files.get('certificate_file')
    if file and file.filename:
        filename = secure_filename(f"cert_{certificate_id}_{file.filename}")
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(upload_path)
        certificate.certificate_file = filename
        db.session.commit()
        flash('Sertifika dosyasÄ± yÃ¼klendi.')
    else:
        flash('Dosya seÃ§ilmedi.')
    return redirect(url_for('admin_certificates'))

@app.route('/certificates')
@login_required
def user_certificates():
    certificates = Certificate.query.filter_by(user_id=current_user.id).all()
    return render_template('user_certificates.html', certificates=certificates)

@app.route('/certificate/<int:certificate_id>')
@login_required
def view_certificate(certificate_id):
    certificate = Certificate.query.get_or_404(certificate_id)
    if certificate.user_id != current_user.id and not current_user.is_admin:
        return redirect(url_for('dashboard'))
    return render_template('view_certificate.html', certificate=certificate)

@app.route('/certificate/<int:certificate_id>/print')
@login_required
def print_certificate(certificate_id):
    certificate = Certificate.query.get_or_404(certificate_id)
    if not current_user.is_admin and current_user.id != certificate.user_id:
        return redirect(url_for('dashboard'))
    return render_template('certificate_print.html', certificate=certificate)

@app.route('/certificate/<int:certificate_id>/download')
@login_required
def download_certificate_pdf(certificate_id):
    """Sertifika sayfasÄ±nÄ± render edip PDF olarak indirir."""
    certificate = Certificate.query.get_or_404(certificate_id)
    # Yetki kontrolÃ¼
    if not current_user.is_admin and current_user.id != certificate.user_id:
        flash('Bu iÅŸlemi yapmaya yetkiniz yok.', 'danger')
        return redirect(url_for('dashboard'))

    # Sertifika ÅŸablonunu render et
    html_string = render_template('view_certificate.html', certificate=certificate, for_pdf=True)
    html = HTML(string=html_string, base_url=request.url_root)
    # PDF olarak dÃ¶ndÃ¼r
    return render_pdf(html, download_filename=f'sertifika_{certificate.certificate_number}.pdf')

@app.route('/admin/generate-certificate/<int:user_id>/<int:course_id>', methods=['POST'])
@login_required
def admin_generate_certificate(user_id, course_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    course = Course.query.get_or_404(course_id)
    
    # KullanÄ±cÄ±nÄ±n tamamladÄ±ÄŸÄ± kurslarÄ± kontrol et
    completed_courses = []
    for c in course.assigned_users:
        if check_course_completion(c, course):
            completed_courses.append(c)
    
    if len(completed_courses) >= course.certificate_type.required_course_count:
        # Belge oluÅŸtur
        certificate = Certificate(
            user_id=user.id,
            certificate_type_id=course.certificate_type_id,
            certificate_number=f"CERT-{uuid.uuid4().hex[:8].upper()}"
        )
        certificate.courses = completed_courses
        db.session.add(certificate)
        db.session.commit()
        flash('Belge baÅŸarÄ±yla oluÅŸturuldu.')
    else:
        flash('KullanÄ±cÄ± belge iÃ§in gerekli kurslarÄ± tamamlamamÄ±ÅŸ.')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/certificates')
@login_required
def admin_certificates():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    certificates = Certificate.query.order_by(Certificate.issue_date.desc()).all()
    return render_template('admin_certificates.html', certificates=certificates)

@app.route('/admin/announcements', methods=['GET', 'POST'])
@login_required
def admin_announcements():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        title = request.form.get('title')
        content = request.form.get('content')
        if title and content:
            ann = Announcement(title=title, content=content)
            db.session.add(ann)
            db.session.commit()
            flash('Duyuru eklendi.', 'success')
        else:
            flash('BaÅŸlÄ±k ve iÃ§erik zorunlu.', 'danger')
    announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
    return render_template('admin_announcements.html', announcements=announcements)

@app.route('/admin/announcements/<int:ann_id>/delete', methods=['POST'])
@login_required
def delete_announcement(ann_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    ann = Announcement.query.get_or_404(ann_id)
    db.session.delete(ann)
    db.session.commit()
    flash('Duyuru silindi.', 'success')
    return redirect(url_for('admin_announcements'))

@app.route('/admin/announcements/<int:ann_id>/edit', methods=['POST'])
@login_required
def edit_announcement(ann_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    ann = Announcement.query.get_or_404(ann_id)
    title = request.form.get('title')
    content = request.form.get('content')
    if title and content:
        ann.title = title
        ann.content = content
        db.session.commit()
        flash('Duyuru gÃ¼ncellendi.', 'success')
    else:
        flash('BaÅŸlÄ±k ve iÃ§erik zorunlu.', 'danger')
    return redirect(url_for('admin_announcements'))

@app.route('/admin/categories', methods=['GET', 'POST'])
@login_required
def admin_categories():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        category = Category(
            name=request.form.get('name'),
            description=request.form.get('description')
        )
        db.session.add(category)
        db.session.commit()
        flash('Kategori baÅŸarÄ±yla eklendi.')
        return redirect(url_for('admin_categories'))
    
    categories = Category.query.all()
    return render_template('admin_categories.html', categories=categories)

@app.route('/admin/reports')
@login_required
def admin_reports():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    courses = Course.query.all()
    return render_template('admin_reports.html', courses=courses, now=datetime.now)

@app.route('/admin/reports/download', methods=['POST'])
@login_required
def download_multi_report():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    course_ids = request.form.getlist('course_ids')
    if not course_ids:
        flash('En az bir kurs seÃ§melisiniz.')
        return redirect(url_for('admin_reports'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Raporlar'
    row = 1
    for course_id in course_ids:
        course = Course.query.get(int(course_id))
        # Kurs baÅŸlÄ±ÄŸÄ±
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value=f"{course.title} - Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=13, color='4f8cff')
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        # SÃ¼tun baÅŸlÄ±klarÄ±
        headers = ['KullanÄ±cÄ± AdÄ±', 'E-posta', 'T. Ä°Ã§erik', 'Tamamlanan', 'Ä°lerleme (%)', 'Test Sonucu', 'Durum', 'Son Tamamlanma']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='6A82FB', end_color='6A82FB', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin', color='B4B4B4'), right=Side(style='thin', color='B4B4B4'), top=Side(style='thin', color='B4B4B4'), bottom=Side(style='thin', color='B4B4B4'))
        row += 1
        # KullanÄ±cÄ± satÄ±rlarÄ±
        for user in course.assigned_users:
            # Yeni progress sistemi ile hesapla
            progress_details = course.get_user_progress(user)
            
            # Test sonucu ve durumu
            test_score = '-'
            test_status = 'Devam Ediyor'
            last_completed_at = ''
            
            if progress_details.passed_test:
                test_score = progress_details.test_score if progress_details.test_score is not None else '-'
                test_status = 'TamamlandÄ±'
            elif progress_details.is_completed:
                test_status = 'TamamlandÄ±'
            
            # Son tamamlanma tarihini bul
            latest_date = None
            for video in course.videos:
                progress = Progress.query.filter_by(user_id=user.id, video_id=video.id).first()
                if progress and progress.completed and progress.completed_at:
                    if latest_date is None or progress.completed_at > latest_date:
                        latest_date = progress.completed_at
            
            if latest_date:
                last_completed_at = latest_date.strftime('%d.%m.%Y %H:%M')
            
            ws.cell(row=row, column=1, value=f"{user.first_name} {user.last_name}")
            ws.cell(row=row, column=2, value=user.email)
            ws.cell(row=row, column=3, value=progress_details.total_steps)
            ws.cell(row=row, column=4, value=progress_details.completed_steps)
            ws.cell(row=row, column=5, value=f"{progress_details.progress_percent}%")
            ws.cell(row=row, column=6, value=test_score)
            ws.cell(row=row, column=7, value=test_status)
            ws.cell(row=row, column=8, value=last_completed_at)
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = Border(left=Side(style='thin', color='B4B4B4'), right=Side(style='thin', color='B4B4B4'), top=Side(style='thin', color='B4B4B4'), bottom=Side(style='thin', color='B4B4B4'))
            row += 1
        # Her kursun tablosundan sonra bir satÄ±r boÅŸluk bÄ±rak
        row += 1
    # SÃ¼tun geniÅŸlikleri (MergedCell hatasÄ±na karÅŸÄ± gÃ¼venli)
    for col in ws.columns:
        first_cell = next((cell for cell in col if hasattr(cell, 'column_letter')), None)
        if not first_cell:
            continue
        column = first_cell.column_letter
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 4
    report_path = os.path.join(app.config['UPLOAD_FOLDER'], f'multi_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    wb.save(report_path)
    return send_file(report_path, as_attachment=True)

@app.route('/course/<int:course_id>/pdf-test', methods=['GET', 'POST'])
@login_required
def pdf_test(course_id):
    course = Course.query.get_or_404(course_id)
    
    # Test daha Ã¶nce tamamlanmÄ±ÅŸ mÄ± kontrol et
    last_video = Video.query.filter_by(course_id=course_id).order_by(Video.order.desc()).first()
    if last_video:
        existing_progress = Progress.query.filter_by(user_id=current_user.id, video_id=last_video.id).first()
        if existing_progress and existing_progress.test_completed:
            # Test tamamlanmÄ±ÅŸsa tekrar giriÅŸi engelle
            if existing_progress.test_score >= course.passing_score:
                flash(f'Bu testi baÅŸarÄ±yla tamamladÄ±nÄ±z! SonuÃ§: %{existing_progress.test_score} (GeÃ§me notu: %{course.passing_score})', 'success')
            else:
                flash(f'Bu testi daha Ã¶nce tamamladÄ±nÄ±z! SonuÃ§: %{existing_progress.test_score} (GeÃ§me notu: %{course.passing_score})', 'warning')
            return redirect(url_for('course', course_id=course_id))
    
    # Test dosyasÄ± tipi ve adÄ±
    test_file_type = None
    test_file_name = None
    if course.test_pdf:
        test_file_type = 'pdf'
        test_file_name = course.test_pdf
    elif course.test_images:
        test_file_type = 'image'
        test_file_name = course.test_images
    if not test_file_type or not course.test_question_count or not course.test_answer_key:
        flash('Bu kurs iÃ§in test tanÄ±mlanmamÄ±ÅŸ.')
        return redirect(url_for('course', course_id=course_id))
    
    # TÃ¼m iÃ§eriÄŸin tamamlanÄ±p tamamlanmadÄ±ÄŸÄ±nÄ± kontrol et (GET isteÄŸi iÃ§in)
    progress_details = course.get_user_progress(current_user)
    if not progress_details.all_content_completed:
        flash('Teste geÃ§mek iÃ§in Ã¶nce tÃ¼m eÄŸitim iÃ§eriÄŸini tamamlamalÄ±sÄ±nÄ±z.', 'warning')
        return redirect(url_for('course', course_id=course_id))

    if request.method == 'POST':
        # TÃ¼m iÃ§eriÄŸin tamamlanÄ±p tamamlanmadÄ±ÄŸÄ±nÄ± kontrol et
        progress_details = course.get_user_progress(current_user)
        if not progress_details.all_content_completed:
            flash('Teste geÃ§mek iÃ§in Ã¶nce tÃ¼m eÄŸitim iÃ§eriÄŸini tamamlamalÄ±sÄ±nÄ±z.', 'warning')
            return redirect(url_for('course', course_id=course_id))
        
        user_answers = []
        for i in range(1, course.test_question_count + 1):
            user_answers.append(request.form.get(f'q{i}', '').upper())
        answer_key = [x.strip() for x in course.test_answer_key.split(',')]
        correct = 0
        for idx, ans in enumerate(user_answers):
            if idx < len(answer_key) and ans == answer_key[idx]:
                correct += 1
        percent = int((correct / course.test_question_count) * 100)
        # Sonucu Progress'e kaydet (son video Ã¼zerinden)
        if last_video:
            progress = Progress.query.filter_by(user_id=current_user.id, video_id=last_video.id).first()
            if not progress:
                progress = Progress(user_id=current_user.id, video_id=last_video.id)
                db.session.add(progress)
            
            # Test daha Ã¶nce tamamlanmÄ±ÅŸsa tekrar kayÄ±t yapma
            if progress.test_completed:
                flash('Bu test daha Ã¶nce tamamlanmÄ±ÅŸ. AynÄ± test tekrar kaydedilemez.', 'warning')
                return redirect(url_for('course', course_id=course_id))
            
            progress.test_score = percent
            progress.test_completed = True  # Test tamamlandÄ± olarak iÅŸaretle
            progress.completed_at = datetime.utcnow()
            db.session.commit()
            
            # BaÅŸarÄ± durumuna gÃ¶re mesaj
            passing_score = course.passing_score or 70
            if percent >= passing_score:
                flash(f'Tebrikler! Testi baÅŸarÄ±yla geÃ§tiniz. SonuÃ§: %{percent} (GeÃ§me notu: %{passing_score})', 'success')
            else:
                flash(f'Test tamamlandÄ± ancak geÃ§eme notunu karÅŸÄ±layamadÄ±nÄ±z. SonuÃ§: %{percent} (GeÃ§me notu: %{passing_score})', 'warning')
        
        return redirect(url_for('course', course_id=course_id))

    return render_template('pdf_test.html', course=course, test_file_type=test_file_type, test_file_name=test_file_name)

@app.route('/admin/certificate-operations', methods=['GET', 'POST'])
@login_required
def admin_certificate_operations():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    courses = Course.query.all()
    certificate_types = CertificateType.query.all()
    course_id = request.args.get('course_id', type=int)
    selected_course = Course.query.get(course_id) if course_id else None
    
    eligible_users = []
    if selected_course:
        for user in selected_course.assigned_users:
            # KullanÄ±cÄ±nÄ±n bu kurs iÃ§in zaten bir sertifikasÄ± olup olmadÄ±ÄŸÄ±nÄ± kontrol et
            existing_cert = Certificate.query.filter(
                Certificate.user_id == user.id,
                Certificate.courses.any(id=selected_course.id)
            ).first()

            if not existing_cert:
                user_progress = selected_course.get_user_progress(user)
                # Sadece kursu tamamlayan kullanÄ±cÄ±larÄ± listeye ekle
                if user_progress.is_completed:
                    eligible_users.append({'user': user, 'progress': user_progress})

    if request.method == 'POST':
        user_ids = request.form.getlist('user_ids')
        certificate_type_id = request.form.get('certificate_type_id')
        certificate_file = request.files.get('certificate_file')

        if not user_ids or not certificate_type_id:
            flash('LÃ¼tfen kullanÄ±cÄ± ve belge tÃ¼rÃ¼ seÃ§in.', 'danger')
            return redirect(url_for('admin_certificate_operations', course_id=course_id))

        filename = None
        if certificate_file and certificate_file.filename:
            # DosyayÄ± gÃ¼venli bir ÅŸekilde kaydet
            filename = secure_filename(f"cert_{datetime.utcnow().timestamp()}_{certificate_file.filename}")
            upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            certificate_file.save(upload_path)
        
        # SeÃ§ilen her kullanÄ±cÄ± iÃ§in sertifika oluÅŸtur
        for user_id in user_ids:
            user = User.query.get(user_id)
            if user:
                certificate = Certificate(
                    user_id=user.id,
                    certificate_type_id=certificate_type_id,
                    certificate_number=f"CERT-{uuid.uuid4().hex[:8].upper()}",
                    certificate_file=filename  # Dosya yoksa None (boÅŸ) olacak
                )
                # SertifikayÄ± kursla iliÅŸkilendir
                if selected_course:
                    certificate.courses.append(selected_course)
                
                db.session.add(certificate)
        
        db.session.commit()
        flash(f'{len(user_ids)} kullanÄ±cÄ±ya baÅŸarÄ±yla sertifika atandÄ±.', 'success')
        return redirect(url_for('admin_certificate_operations', course_id=course_id))
    
    return render_template('admin_certificate_operations.html', 
                           courses=courses, 
                           selected_course=selected_course,
                           eligible_users=eligible_users,
                           certificate_types=certificate_types,
                           course_id=course_id)

@app.route('/admin/user-certificates', methods=['GET', 'POST'])
@login_required
def admin_user_certificates():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    users = User.query.filter_by(is_admin=False).all()
    selected_user_id = request.args.get('user_id', type=int)
    selected_user = User.query.get(selected_user_id) if selected_user_id else None
    user_certificates = []
    if selected_user:
        user_certificates = Certificate.query.filter_by(user_id=selected_user.id).all()
    return render_template('admin_user_certificates.html', users=users, selected_user=selected_user, user_certificates=user_certificates)

@app.route('/admin/certificate/<int:certificate_id>/delete', methods=['POST'])
@login_required
def admin_delete_certificate(certificate_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    cert = Certificate.query.get_or_404(certificate_id)
    
    # Sertifika dosyasÄ±nÄ± sil
    if cert.certificate_file:
        cert_path = os.path.join(app.config['UPLOAD_FOLDER'], cert.certificate_file)
        if os.path.exists(cert_path):
            try:
                os.remove(cert_path)
                print(f"Sertifika dosyasÄ± silindi: {cert_path}")
            except Exception as e:
                print(f"Sertifika dosyasÄ± silinemedi: {cert_path} - {e}")
    
    user_id = cert.user_id
    db.session.delete(cert)
    db.session.commit()
    flash('Sertifika ve dosyasÄ± baÅŸarÄ±yla silindi.')
    return redirect(url_for('admin_user_certificates', user_id=user_id))

@app.route('/admin/database')
@login_required
def admin_database():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    from models import User, Course, Video, Progress, Category, CertificateType, Certificate, Announcement
    
    # Database istatistikleri
    stats = {
        'users': User.query.count(),
        'courses': Course.query.count(),
        'videos': Video.query.count(),
        'categories': Category.query.count(),
        'certificate_types': CertificateType.query.count(),
        'certificates': Certificate.query.count(),
        'announcements': Announcement.query.count(),
        'total_progress': Progress.query.count()
    }
    
    # Son eklenen kayÄ±tlar
    recent_users = User.query.order_by(User.created_at.desc()).limit(5).all()
    recent_courses = Course.query.order_by(Course.created_at.desc()).limit(5).all()
    
    return render_template('admin_database.html', stats=stats, recent_users=recent_users, recent_courses=recent_courses)

@app.route('/admin/database/backup', methods=['POST'])
@login_required
def backup_database():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    import shutil
    from datetime import datetime
    
    try:
        # Mevcut database dosyasÄ±
        source_db = 'instance/isg.db'
        if not os.path.exists(source_db):
            flash('Database dosyasÄ± bulunamadÄ±!', 'danger')
            return redirect(url_for('admin_database'))
        
        # Yedek dosya adÄ±
        backup_name = f'isg_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
        backup_path = os.path.join('instance', backup_name)
        
        # Database'i kopyala
        shutil.copy2(source_db, backup_path)
        
        flash(f'Database baÅŸarÄ±yla yedeklendi: {backup_name}', 'success')
        
    except Exception as e:
        flash(f'Yedekleme hatasÄ±: {str(e)}', 'danger')
    
    return redirect(url_for('admin_database'))

@app.route('/admin/database/reset', methods=['POST'])
@login_required
def reset_database():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    try:
        # Database'i sÄ±fÄ±rla
        from create_db import create_db
        create_db()
        flash('Database baÅŸarÄ±yla sÄ±fÄ±rlandÄ±!', 'success')
        
    except Exception as e:
        flash(f'Database sÄ±fÄ±rlama hatasÄ±: {str(e)}', 'danger')
    
    return redirect(url_for('admin_database'))

@app.route('/admin/database/export', methods=['POST'])
@login_required
def export_database():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    try:
        import sqlite3
        import json
        from datetime import datetime
        
        # Database baÄŸlantÄ±sÄ±
        conn = sqlite3.connect('instance/isg.db')
        cursor = conn.cursor()
        
        # TÃ¼m tablolarÄ± al
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        
        export_data = {}
        
        for table in tables:
            table_name = table[0]
            cursor.execute(f"SELECT * FROM {table_name}")
            rows = cursor.fetchall()
            
            # SÃ¼tun adlarÄ±nÄ± al
            cursor.execute(f"PRAGMA table_info({table_name});")
            columns = [column[1] for column in cursor.fetchall()]
            
            # Verileri dictionary formatÄ±nda sakla
            table_data = []
            for row in rows:
                row_dict = {}
                for i, value in enumerate(row):
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    row_dict[columns[i]] = value
                table_data.append(row_dict)
            
            export_data[table_name] = table_data
        
        conn.close()
        
        # JSON dosyasÄ± olarak kaydet
        export_filename = f'database_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        export_path = os.path.join('static', 'uploads', export_filename)
        
        with open(export_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
        
        return send_file(export_path, as_attachment=True, download_name=export_filename)
        
    except Exception as e:
        flash(f'Export hatasÄ±: {str(e)}', 'danger')
        return redirect(url_for('admin_database'))

@app.route('/admin/database/import', methods=['POST'])
@login_required
def import_database():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    if 'database_file' not in request.files:
        flash('Dosya seÃ§ilmedi!', 'danger')
        return redirect(url_for('admin_database'))
    
    file = request.files['database_file']
    if file.filename == '':
        flash('Dosya seÃ§ilmedi!', 'danger')
        return redirect(url_for('admin_database'))
    
    if not file.filename.endswith('.json'):
        flash('Sadece JSON dosyalarÄ± kabul edilir!', 'danger')
        return redirect(url_for('admin_database'))
    
    try:
        import json
        import sqlite3
        
        # JSON dosyasÄ±nÄ± oku
        data = json.load(file)
        
        # Database baÄŸlantÄ±sÄ±
        conn = sqlite3.connect('instance/isg.db')
        cursor = conn.cursor()
        
        # Mevcut verileri temizle
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        
        for table in tables:
            table_name = table[0]
            cursor.execute(f"DELETE FROM {table_name}")
        
        # Yeni verileri ekle
        for table_name, table_data in data.items():
            if table_data:
                # Ä°lk satÄ±rdan sÃ¼tun adlarÄ±nÄ± al
                columns = list(table_data[0].keys())
                placeholders = ', '.join(['?' for _ in columns])
                column_names = ', '.join(columns)
                
                for row in table_data:
                    values = [row.get(col) for col in columns]
                    cursor.execute(f"INSERT INTO {table_name} ({column_names}) VALUES ({placeholders})", values)
        
        conn.commit()
        conn.close()
        
        flash('Database baÅŸarÄ±yla import edildi!', 'success')
        
    except Exception as e:
        flash(f'Import hatasÄ±: {str(e)}', 'danger')
    
    return redirect(url_for('admin_database'))

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'message': 'E-posta adresi gerekli.', 'category': 'danger'}), 400

        user = User.query.filter_by(email=email).first()
        
        if user:
            # Aktif (onaylanmÄ±ÅŸ veya bekleyen) bir istek var mÄ± kontrol et
            existing_request = PasswordReset.query.filter(
                PasswordReset.user_id == user.id,
                PasswordReset.status.in_(['pending', 'approved'])
            ).first()

            if existing_request:
                # EÄŸer istek ONAYLANMIÅSA ve token geÃ§erliyse, yenileme sayfasÄ±na yÃ¶nlendir
                if existing_request.status == 'approved' and existing_request.token and existing_request.expires_at > datetime.utcnow():
                    return jsonify({
                        'redirect_url': url_for('reset_password_with_token', token=existing_request.token)
                    })
                
                # EÄŸer istek BEKLEMEDEYSE, bilgi ver
                return jsonify({
                    'message_long': get_status_message('pending', True), 
                    'category': 'info'
                })

            # Aktif istek yoksa yeni bir tane oluÅŸtur
            reset_request = PasswordReset(user_id=user.id)
            db.session.add(reset_request)
            db.session.commit()
            
            return jsonify({
                'message_long': get_status_message('pending', True),
                'category': 'success'
            })
        else:
            return jsonify({'message': 'Bu e-posta adresi ile kayÄ±tlÄ± kullanÄ±cÄ± bulunamadÄ±.', 'category': 'danger'}), 404
    
    return render_template('forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password_with_token(token):
    # Token'a gÃ¶re isteÄŸi bul
    reset_request = PasswordReset.query.filter_by(token=token, status='approved').first()

    # Token geÃ§erli deÄŸilse veya sÃ¼resi dolmuÅŸsa
    if not reset_request or reset_request.expires_at < datetime.utcnow():
        flash('Åifre sÄ±fÄ±rlama linki geÃ§ersiz veya sÃ¼resi dolmuÅŸ. LÃ¼tfen yeni bir istek gÃ¶nderin.', 'danger')
        return redirect(url_for('forgot_password'))

    form = ResetPasswordForm()
    if form.validate_on_submit():
        # Åifreyi gÃ¼ncelle ve isteÄŸi tamamla
        user = reset_request.user
        user.set_password(form.password.data)
        reset_request.status = 'completed'
        db.session.commit()

        flash('Åifreniz baÅŸarÄ±yla gÃ¼ncellendi! Yeni ÅŸifrenizle giriÅŸ yapabilirsiniz.', 'success')
        return redirect(url_for('login'))

    return render_template('reset_password_with_token.html', token=token, form=form)

@app.route('/admin/password-requests')
@login_required
def admin_password_requests():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    pending_requests = PasswordReset.query.filter_by(status='pending').order_by(PasswordReset.requested_at.desc()).all()
    
    # OnaylanmÄ±ÅŸ, tamamlanmÄ±ÅŸ ve reddedilmiÅŸ olanlar
    processed_requests = PasswordReset.query.filter(PasswordReset.status != 'pending').order_by(PasswordReset.approved_at.desc()).limit(20).all()
    
    return render_template('admin_password_requests.html', 
                         pending_requests=pending_requests,
                         processed_requests=processed_requests)

@app.route('/admin/password-request/<int:request_id>/approve', methods=['POST'])
@login_required
def approve_password_request(request_id):
    if not current_user.is_admin:
        flash('Yetkisiz iÅŸlem.', 'danger')
        return redirect(url_for('dashboard'))
    
    reset_request = PasswordReset.query.get_or_404(request_id)
    
    # Token oluÅŸtur ve son kullanma tarihi belirle (Ã¶rn: 1 saat)
    reset_request.token = secrets.token_urlsafe(32)
    reset_request.expires_at = datetime.utcnow() + timedelta(hours=1)
    reset_request.status = 'approved'
    reset_request.approved_by = current_user.id
    reset_request.approved_at = datetime.utcnow()
    
    db.session.commit()
    flash(f'{reset_request.user.username} kullanÄ±cÄ±sÄ±nÄ±n isteÄŸi onaylandÄ±. KullanÄ±cÄ± artÄ±k ÅŸifresini sÄ±fÄ±rlayabilir.', 'success')

    return redirect(url_for('admin_password_requests'))

@app.route('/admin/password-request/<int:request_id>/reject', methods=['POST'])
@login_required
def reject_password_request(request_id):
    if not current_user.is_admin:
        flash('Yetkisiz iÅŸlem.', 'danger')
        return redirect(url_for('dashboard'))
    
    reset_request = PasswordReset.query.get_or_404(request_id)
    reset_request.status = 'rejected'
    reset_request.approved_by = current_user.id
    reset_request.approved_at = datetime.utcnow()
    db.session.commit()
    
    flash(f'{reset_request.user.username} kullanÄ±cÄ±sÄ±nÄ±n ÅŸifre sÄ±fÄ±rlama isteÄŸi reddedildi.', 'warning')
    return redirect(url_for('admin_password_requests'))

@app.route('/check-password-status')
def check_password_status():
    email = request.args.get('email')
    if not email:
        return jsonify({'status': 'error', 'message': 'E-posta adresi gerekli'})
    
    user = User.query.filter_by(email=email).first()
    if not user:
        return jsonify({'status': 'error', 'message': 'KullanÄ±cÄ± bulunamadÄ±'})
    
    reset_request = PasswordReset.query.filter(
        PasswordReset.user_id == user.id,
        PasswordReset.status.in_(['pending', 'approved'])
    ).order_by(PasswordReset.requested_at.desc()).first()
    
    if not reset_request:
        return jsonify({'status': 'none', 'message': 'Aktif bir ÅŸifre sÄ±fÄ±rlama isteÄŸi bulunamadÄ±.'})
    
    response_data = {
        'status': reset_request.status,
        'message_long': get_status_message(reset_request.status, True),
        'requested_at': reset_request.requested_at.strftime('%d.%m.%Y %H:%M')
    }

    if reset_request.status == 'approved':
        response_data['token'] = reset_request.token

    return jsonify(response_data)

def get_status_message(status, long=False):
    messages = {
        'pending': 'Ä°steÄŸiniz admin onayÄ± iÃ§in gÃ¶nderildi. LÃ¼tfen daha sonra tekrar kontrol edin.',
        'approved': 'Ä°steÄŸiniz onaylandÄ±! Åifrenizi ÅŸimdi sÄ±fÄ±rlayabilirsiniz.',
        'rejected': 'Ä°steÄŸiniz reddedildi. LÃ¼tfen admin ile iletiÅŸime geÃ§in.',
        'completed': 'Åifreniz gÃ¼ncellendi. Yeni ÅŸifrenizle giriÅŸ yapabilirsiniz.'
    }
    long_messages = {
        'pending': """
            <strong>Ä°steÄŸiniz AlÄ±ndÄ± ve Admin OnayÄ±na GÃ¶nderildi</strong><br>
            <small>
            Bir yÃ¶neticinin isteÄŸinizi onaylamasÄ± gerekiyor. OnaylandÄ±ktan sonra bu sayfaya tekrar gelerek ÅŸifrenizi yenileyebilirsiniz.
            </small>
        """,
        'approved': """
            <strong>Ä°steÄŸiniz OnaylandÄ±!</strong><br>
            <small>
            Åimdi ÅŸifrenizi sÄ±fÄ±rlayabilirsiniz. LÃ¼tfen aÅŸaÄŸÄ±daki butonu kullanarak yeni ÅŸifrenizi belirleyin. Bu link 1 saat geÃ§erlidir.
            </small>
        """,
        'rejected': """
            <strong>Ä°steÄŸiniz Reddedildi</strong><br>
            <small>
            Åifre sÄ±fÄ±rlama isteÄŸiniz bir admin tarafÄ±ndan reddedildi. 
            Daha fazla bilgi almak veya nedenini Ã¶ÄŸrenmek iÃ§in lÃ¼tfen kurumunuzun yetkilisiyle iletiÅŸime geÃ§in.
            </small>
        """,
        'completed': """
            <strong>Åifreniz Zaten GÃ¼ncellendi</strong><br>
            <small>
            Bu istek kullanÄ±larak ÅŸifreniz daha Ã¶nce gÃ¼ncellenmiÅŸ. Yeni ÅŸifrenizle giriÅŸ yapabilirsiniz. 
            Tekrar sÄ±fÄ±rlamak isterseniz, yeni bir istek oluÅŸturun.
            </small>
        """
    }
    if long:
        return long_messages.get(status, 'Bilinmeyen durum.')
    return messages.get(status, 'Bilinmeyen durum')

@app.route('/admin/manage-course/<int:course_id>', methods=['GET', 'POST'])
@login_required
def admin_manage_course(course_id):
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    
    course = Course.query.get_or_404(course_id)
    categories = Category.query.all()
    
    if request.method == 'POST':
        # Sadece kurs bilgilerini gÃ¼ncelle
        course.title = request.form.get('title')
        course.description = request.form.get('description')
        course.category_id = request.form.get('category_id')
        course.passing_score = request.form.get('passing_score', 70, type=int)
        course.test_required = 'test_required' in request.form
        
        db.session.commit()
        flash('Kurs bilgileri baÅŸarÄ±yla gÃ¼ncellendi.', 'success')
        return redirect(url_for('admin_manage_course', course_id=course_id))
    
    return render_template('admin_manage_course.html', course=course, categories=categories)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(host='0.0.0.0', port=5000, debug=True)